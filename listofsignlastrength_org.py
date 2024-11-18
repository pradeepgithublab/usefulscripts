import requests
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.styles import Font

# Replace with your Mist API credentials
MIST_API_TOKEN = 'xh9oBdbYQllXCmCh5Ad63h3Y3enrewPWrDD37XpYA5f1ealYTPJQEb2FMkmo0DBi9vFIiLJpNlHd6fM95Zog8e0NldKTnOol'
ORG_ID = '15e7597e-9b06-4381-8443-16aba95c5e0d'
BASE_URL = 'https://api.eu.mist.com/api/v1'

# Set up headers for requests
headers = {
    'Authorization': f'Token {MIST_API_TOKEN}'
}

# Function to get all sites in the organization
def get_sites():
    url = f'{BASE_URL}/orgs/{ORG_ID}/sites'
    response = requests.get(url, headers=headers)
    
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Error fetching sites: {response.status_code}, {response.text}")
        return []

# Function to get live users for a specific site and filter by signal strength
def get_live_users_for_site(site):
    site_id = site['id']
    site_name = site['name']
    url = f'{BASE_URL}/sites/{site_id}/stats/clients'
    
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        clients = response.json()

        # Filter clients by signal strength (RSSI) >= -70 dBm
        filtered_clients = [
            {
                'Site Name': site_name,
                'MAC Address': client.get('mac'),
                'SSID': client.get('ssid', 'N/A'),
                'Signal Strength (dBm)': client.get('rssi', -100)
            }
            for client in clients
            if client.get('rssi', -100) >= -70
        ]
        
        return filtered_clients
    except requests.RequestException as e:
        print(f"Error fetching clients for site {site_name}: {e}")
        return []

# Function to get client details across the entire organization in parallel
def get_all_clients_parallel():
    all_clients = []
    sites = get_sites()

    # Use ThreadPoolExecutor to fetch data from multiple sites concurrently
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {executor.submit(get_live_users_for_site, site): site for site in sites}
        
        for future in as_completed(futures):
            site_clients = future.result()
            all_clients.extend(site_clients)

    return all_clients

# Function to save data to Excel and apply formatting
def save_to_excel_with_formatting(clients_data):
    # Convert to DataFrame
    df = pd.DataFrame(clients_data)
    
    # Save DataFrame to Excel
    file_path = 'Org_Client_Signal_Strength.xlsx'
    df.to_excel(file_path, index=False, sheet_name='Client Data')
    
    # Load workbook and apply formatting for signal strength
    workbook = load_workbook(file_path)
    sheet = workbook['Client Data']
    
    # Define red font for low signal strength
    red_font = Font(color="FF0000")  # Red for signal strength below -70 dBm
    
    # Apply conditional formatting
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=4, max_col=4):  # Column 4 is Signal Strength (dBm)
        cell = row[0]
        if cell.value is not None and cell.value < -70:
            cell.font = red_font

    # Save the updated workbook
    workbook.save(file_path)
    print("Data has been saved with color formatting to Org_Client_Signal_Strength.xlsx")

if __name__ == "__main__":
    clients_data = get_all_clients_parallel()
    if clients_data:
        save_to_excel_with_formatting(clients_data)
    else:
        print("No client data to save.")
